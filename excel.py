from openpyxl import Workbook, load_workbook
from word import *
import os
import matplotlib.pyplot as plt
import numpy as np
from scipy.stats import ttest_1samp, ttest_ind, pearsonr

def write_sense_list(word):

    if word.length == 0:
        print("Sense List not possible for given input: " + word.label)
        exit()

    wb = Workbook()
    ws1 = wb.active

    ws1['A1'], ws1['B1'], ws1['C1'], ws1['D1'], ws1['E1'] = "Starting Date", "Word Form", "Identifiers", "Categories", "PoS"
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 110
    ws1.column_dimensions['D'].width = 35
    for num in range(len(word.senses)):
        ws1.cell(row=num+2, column=1).value = word.senses[num].date
        ws1.cell(row=num+2, column=2).value = word.senses[num].word_form
        ws1.cell(row=num+2, column=3).value = word.senses[num].identifiers
        ws1.cell(row=num+2, column=4).value = word.senses[num].categories
        ws1.cell(row=num+2, column=5).value = word.senses[num].PoS

    save_loc =  os.getcwd() + "\\Sense Lists\\" + "[" + word.PoS + "] " + word.label + ".xlsx"
    wb.save(save_loc)
    return save_loc

def write_sense_matrix(word):

    if word.length == 0 or word.length == 1:
        print("Sense Matrix not possible for given input: " + word.label)
        exit()

    wb = Workbook()
    ws = wb.active
    length = len(word.senses)

    for num in range(2, length+2):
        r, c = ws.cell(row=1, column=num), ws.cell(row=num, column=1) #1st row and 1st column
        r.value, c.value = str(word.senses[num-2]), str(word.senses[num-2])

    for row in range(2, length+2):
        for col in range(2, length+2):
            cur_cell = ws.cell(row=row, column=col)
            # when a sense is matched with itself, its difference value is maximal
            if row == col:
                cur_cell.value = 0
            else:
                similarity_score = 0
                list_range = min(len(word.senses[row-2].listed_cat), len(word.senses[col-2].listed_cat))
                for r in range(list_range):
                    if word.senses[row-2].listed_cat[r] == word.senses[col-2].listed_cat[r]:
                        similarity_score += 1
                    else:
                        break
                cur_cell.value = exp(-similarity_score)


    save_loc =  os.getcwd() + "\\Sense Matrices\\" + "[" + word.PoS + "] " + word.label + ".xlsx"
    wb.save(save_loc)
    return save_loc

def write_baseline_comparison_chart(word, worksheet):

    if not word.viable:
        print("Chart not possible for given input: " + word.label)

    for i in range(len(word.vs_baseline)):
        create_sense_by_n(i*2+1, i, word, worksheet)

    for j in range(2, len(word.nb_senses)+2):
        worksheet.cell(row=i*2+3, column=j).value = word.nb_senses[j-2].date
        worksheet.cell(row=i*2+4, column=j).value = word.avgs_by_date[j-2]

    worksheet.cell(row=i*2+3, column=j+1).value = "avg_before R"
    worksheet.cell(row=i*2+4, column=j+1).value = word.pearson_avg_before

    worksheet.cell(row=i*2+3, column=j+2).value = "avg_after R"
    worksheet.cell(row=i*2+4, column=j+2).value = word.pearson_avg_after

    save_loc = os.getcwd() + "\\Comparison Charts\\" + "[" + word.PoS + "] " + word.label + ".xlsx"

    return save_loc

def write_coca_iteration(part_of_speech, num_words):
    cj_start, cj_end = 14, 51
    p_start, p_end = 87, 183
    aj_start, aj_end = 184, 1022
    n_start, n_end = 1058, 3599
    av_start, av_end = 3646, 3985
    v_start, v_end = 4000, 5000
    coca = load_workbook(filename = 'top_5000_coca_words.xlsx')
    ws = coca.active
    word_selection = []

    if part_of_speech == "cj":
        if num_words > (cj_end - cj_start):
            print(str(num_words) + " is too large. Creating maximal amount at: " + str(cj_end - cj_start + 1) + " words.")
        start, end = cj_start, min(cj_end, cj_start + num_words - 1)

    if part_of_speech == "p":
        if num_words > (p_end - p_start):
            print(str(num_words)+ " is too large. Creating maximal amount at: " + str(p_end - p_start + 1) + " words.")
        start, end = p_start, min(p_end, p_start + num_words - 1)

    if part_of_speech == "aj":
        if num_words > (aj_end - aj_start):
            print(str(num_words) + " is too large. Creating maximal amount at: " + str(aj_end - aj_start + 1) + " words.")
        start, end = aj_start, min(aj_end, aj_start + num_words - 1)

    if part_of_speech == "n":
        if num_words > (n_end - n_start):
            print(str(num_words) + " is too large. Creating maximal amount at: " + str(n_end - n_start + 1) + " words.")
        start, end = n_start, min(n_end, n_start + num_words - 1)

    if part_of_speech == "av":
        if num_words > (av_end - av_start):
            print(str(num_words) + " is too large. Creating maximal amount at: " + str(av_end - av_start + 1) + " words.")
        start, end = av_start, min(av_end, av_start + num_words - 1)

    if part_of_speech == "v":
        if num_words > (v_end - v_start):
            print(str(num_words) + " is too large. Creating maximal amount at: " + str(v_end - v_start + 1) + " words.")
        start, end = v_start, min(v_end, v_start + num_words - 1)

    for i in range(start, end + 1):
        word_selection.append(ws.cell(row=i, column=2).value[3::])

    word_objects = []
    defunct_objects = []

    for word in word_selection:
        curr_word = Word(word, part_of_speech)
        if curr_word.viable:
            word_objects.append(curr_word)
        else:
            defunct_objects.append(curr_word)
            print(curr_word.label + " not appended due to computational errors or lack of non-baseline senses.", flush=True)

    print("Non-viable words: ")
    print([word.label for word in defunct_objects])

    sorted_words = sorted(zip(word_objects, [word.pearson_avg_after for word in word_objects]),key=lambda x: x[1])    #sorts by the pearson_avg_after of each word
    sorted_words, pearsons = [word[0] for word in sorted_words], [word[1] for word in sorted_words]                   #back to a normal word objects list, but sorted

    wb = Workbook()
    print("Begin writing multiple baseline comparison charts.")
    write_multiple_charts(sorted_words, wb)
    save_loc = os.getcwd() + "\\COCA Corpus Iterations\\" + "top_" + str(num_words) + "_viable_" + part_of_speech + ".xlsx"
    hist_save_loc = os.getcwd() + "\\COCA Corpus Iterations\\Histograms\\" + "top_" + str(num_words) + "_viable_" + part_of_speech + ".pdf"
    wb.save(save_loc)

    write_histogram(pearsons, "COCA - Histogram of " + str(len(pearsons)) + " Sense Correlations " + "[" + part_of_speech + "]"
                   , hist_save_loc)

def write_bnc_iteration(part_of_speech, num_words):
    cj_start, cj_end = 1553, 1586
    p_start, p_end = 4922, 4992
    aj_start, aj_end = 2, 1125
    n_start, n_end = 1660, 4921
    av_start, av_end = 1126, 1552
    v_start, v_end = 5039, 6319
    bnc = load_workbook(filename = 'top_6318_bnc_words.xlsx')
    ws = bnc.active
    word_selection = []

    if part_of_speech == "cj":
        if num_words > (cj_end - cj_start):
            print(str(num_words) + " is too large. Creating maximal amount at: " + str(cj_end - cj_start + 1) + " words.")
        start, end = cj_start, min(cj_end, cj_start + num_words - 1)

    if part_of_speech == "p":
        if num_words > (p_end - p_start):
            print(str(num_words)+ " is too large. Creating maximal amount at: " + str(p_end - p_start + 1) + " words.")
        start, end = p_start, min(p_end, p_start + num_words - 1)

    if part_of_speech == "aj":
        if num_words > (aj_end - aj_start):
            print(str(num_words) + " is too large. Creating maximal amount at: " + str(aj_end - aj_start + 1) + " words.")
        start, end = aj_start, min(aj_end, aj_start + num_words - 1)

    if part_of_speech == "n":
        if num_words > (n_end - n_start):
            print(str(num_words) + " is too large. Creating maximal amount at: " + str(n_end - n_start + 1) + " words.")
        start, end = n_start, min(n_end, n_start + num_words - 1)

    if part_of_speech == "av":
        if num_words > (av_end - av_start):
            print(str(num_words) + " is too large. Creating maximal amount at: " + str(av_end - av_start + 1) + " words.")
        start, end = av_start, min(av_end, av_start + num_words - 1)

    if part_of_speech == "v":
        if num_words > (v_end - v_start):
            print(str(num_words) + " is too large. Creating maximal amount at: " + str(v_end - v_start + 1) + " words.")
        start, end = v_start, min(v_end, v_start + num_words - 1)

    for i in range(start, end + 1):
        word_selection.append(ws.cell(row=i, column=2).value)

    word_objects = []
    defunct_objects = []

    for word in word_selection:
        curr_word = Word(word, part_of_speech)
        if curr_word.viable:
            word_objects.append(curr_word)
        else:
            defunct_objects.append(curr_word)
            print( curr_word.label + " not appended due to computational errors or lack of non-baseline senses.", flush=True)

    print("Non-viable words: ")
    print([word.label for word in defunct_objects])

    sorted_words = sorted(zip(word_objects, [word.pearson_avg_after for word in word_objects]),key=lambda x: x[1])    #sorts by the pearson_avg_after of each word
    sorted_words, pearsons = [word[0] for word in sorted_words], [word[1] for word in sorted_words]                   #back to a normal word objects list, but sorted

    wb = Workbook()
    print("Begin writing multiple baseline comparison charts.")
    write_multiple_charts(sorted_words, wb)
    save_loc = os.getcwd() + "\\BNC Corpus Iterations\\" + "top_" + str(num_words) + "_viable_" + part_of_speech + ".xlsx"
    hist_save_loc = os.getcwd() + "\\BNC Corpus Iterations\\Histograms\\" + "top_" + str(num_words) + "_viable_" + part_of_speech + ".pdf"
    wb.save(save_loc)

    write_histogram(pearsons, "BNC - Histogram of " + str(len(pearsons)) + " Sense Correlations " + "[" + part_of_speech + "]",
                    hist_save_loc)

def write_histogram(correlations, title, save_loc):
    values, bins, patches = plt.hist(correlations, histtype="bar", bins=25)
    plt.xlabel('Correlations')
    plt.ylabel('Counts')
    plt.title(title)
    plt.grid(True)
    plt.savefig(save_loc)
    plt.clf()

def write_multiple_charts(words, workbook):
    worksheets = {}
    pearson_ws = workbook.create_sheet(title="Pearson")

    for i in range(len(words)):
        worksheets["ws" + str(i)] = workbook.create_sheet(title=words[i].label)

    pearson_ws.cell(row=1, column=1).value = "Word"
    pearson_ws.cell(row=1, column=2).value = "Correlation (Avg After)"
    pearson_ws.cell(row=1, column=3).value, pearson_ws.cell(row=2, column=3).value = "Total Words", len(words)
    pearson_ws.cell(row=3, column=3).value, pearson_ws.cell(row=4, column=3).value = "Positive", len([word for word in words if word.pearson_avg_after > 0])
    pearson_ws.cell(row=5, column=3).value, pearson_ws.cell(row=6, column=3).value = "Negative", len([word for word in words if word.pearson_avg_after < 0])
    pearson_ws.cell(row=7, column=3).value, pearson_ws.cell(row=8, column=3).value = "Zero", len([word for word in words if word.pearson_avg_after == 0])

    for j in range(len(words)):
        current_ws = worksheets["ws" + str(j)]
        write_baseline_comparison_chart(words[j], current_ws)
        print("Chart written for word: " + words[j].label, flush=True)
        pearson_ws.cell(row=j + 2, column=1).value = words[j].label
        pearson_ws.cell(row=j + 2, column=2).value = words[j].pearson_avg_after

def create_sense_by_n(input_row, baseline_index, word, worksheet):
    vs = word.vs_baseline[baseline_index]
    length = len(vs.nb_senses)
    base_cell = worksheet.cell(row=input_row+1, column=1)
    base_cell.value = str(vs.bl_sense)

    for col in range(2, length+2):
        worksheet.cell(row=input_row, column=col).value = vs.nb_senses[col-2].date
        worksheet.cell(row=input_row+1, column=col).value = vs.sim_scores[col-2]

    worksheet.cell(row=input_row, column=length+2).value = "Pearson R"
    worksheet.cell(row=input_row+1, column=length+2).value = vs.pearsonR

def write_definitive_bnc():

    b_pear, b_freq, b_aext, b_nums, b_len = [], [], [], [], []
    bnc = load_workbook(filename='BNC - Noun, Verb, Adj, Adv only - Valid Only - Order Sort.xlsx')
    ws = bnc.active

    fig, axes = plt.subplots(nrows=2, ncols=5)
    ax0, ax1, ax2, ax3, ax4, ax5, ax6, ax7, ax8, ax9 = axes.flat

    for i in range(2, 502):
        b_pear.append(ws.cell(row=i, column=10).value)
        b_freq.append(np.log(int(ws.cell(row=i, column=4).value)))
        b_aext.append(ws.cell(row=i, column=11).value)
        b_nums.append(ws.cell(row=i, column=12).value)
        b_len.append(ws.cell(row=i, column=2).value)

    b_len = [len(word) for word in b_len]

    b_conc, b_concp = [], []
    bnc = load_workbook(filename='BNC - Noun, Verb, Adj, Adv only - Valid Only - With Concrete.xlsx')
    ws = bnc.active

    for i in range(2, 502):
        b_conc.append(ws.cell(row=i, column=8).value)
        b_concp.append(ws.cell(row=i, column=9).value)

    def t_test(data):
        stat = round(ttest_1samp(data, 0)[0], 5)
        p = round(ttest_1samp(data, 0)[1], 5)
        return stat, p

    #Histogram of Pearsons

    ax0.hist(b_pear, 25, range=(-1,1))
    ax0.axvline(np.mean(b_pear), color='r', linestyle='dashed', linewidth=2)
    ax0.axvline(np.median(b_pear), color='r', linewidth=2)
    ax0.set_title('Pearson Histogram - T-Stat, P-Value: \n' + str(t_test(b_pear)))

    #Scatterplots:

    write_scatterplot(b_freq, b_pear, ax1, "Log(Freq) vs. Pearson")
    write_scatterplot(b_freq, b_nums, ax2, "Log(Freq) vs. Num Senses")
    write_scatterplot(b_freq, b_aext, ax3, "Log(Freq) vs. Externality")
    write_scatterplot(b_aext, b_pear, ax4, "Externality vs. Pearson")
    write_scatterplot(b_aext, b_nums, ax5, "Externality vs. Num Senses")
    write_scatterplot(b_len, b_nums, ax6, "Length vs. Num Senses")
    write_scatterplot(b_len, b_aext, ax7, "Length vs. Externality")
    write_scatterplot(b_conc, b_concp, ax8, "Concreteness vs. Pearson")

    fig.subplots_adjust(hspace=.5)
    plt.show()

def write_definitive_swadesh():

    s_pear, s_freq, s_aext, s_nums, s_len = [], [], [], [], []
    bnc = load_workbook(filename='Swadesh List.xlsx')
    ws = bnc.active

    fig, axes = plt.subplots(nrows=2, ncols=5)
    ax0, ax1, ax2, ax3, ax4, ax5, ax6, ax7, ax8, ax9 = axes.flat

    for i in range(2, 201):
        if ws.cell(row=i, column=2).value and ws.cell(row=i, column=4).value != "Not Viable":
            s_freq.append(np.log(int(ws.cell(row=i, column=2).value)))
            s_pear.append(ws.cell(row=i, column=3).value)
            s_aext.append(ws.cell(row=i, column=4).value)
            s_nums.append(ws.cell(row=i, column=5).value)
            s_len.append(ws.cell(row=i, column=1).value)

    s_len = [len(word) for word in s_len]

    #s_conc, s_concp = [], []
    bnc = load_workbook(filename='Swadesh List.xlsx')
    ws = bnc.active

    """for i in range(2, 502):
        s_conc.append(ws.cell(row=i, column=8).value)
        s_concp.append(ws.cell(row=i, column=9).value)"""

    def t_test(data):
        stat = round(ttest_1samp(data, 0)[0], 5)
        p = round(ttest_1samp(data, 0)[1], 5)
        return stat, p

    #Histogram of Pearsons

    ax0.hist(s_pear, 25, range=(-1, 1))
    ax0.axvline(np.mean(s_pear), color='r', linestyle='dashed', linewidth=2)
    ax0.axvline(np.median(s_pear), color='r', linewidth=2)
    ax0.set_title('Pearson Histogram - T-Stat, P-Value: \n' + str(t_test(s_pear)))

    #Scatterplots:

    write_scatterplot(s_freq, s_pear, ax1, "Log(Freq) vs. Pearson")
    write_scatterplot(s_freq, s_nums, ax2, "Log(Freq) vs. Num Senses")
    write_scatterplot(s_freq, s_aext, ax3, "Log(Freq) vs. Externality")
    write_scatterplot(s_aext, s_pear, ax4, "Externality vs. Pearson")
    write_scatterplot(s_aext, s_nums, ax5, "Externality vs. Num Senses")
    write_scatterplot(s_len, s_nums, ax6, "Length vs. Num Senses")
    write_scatterplot(s_len, s_aext, ax7, "Length vs. Externality")
    #write_scatterplot(s_conc, s_concp, ax8, "Concreteness vs. Pearson")

    fig.subplots_adjust(hspace=.5)
    plt.show()

def write_5_by_2_std_plot(before, limited):

    if before:
        b_col, c_col = 5, 6
    else:
        b_col, c_col = 6, 7

    c_aj_start, c_aj_end = 2, 580
    c_n_start, c_n_end = 581, 2640
    c_av_start, c_av_end = 2641, 2809
    c_v_start, c_v_end = 2810, 3724
    if limited:
        c_aj_end, c_n_end, c_av_end, c_v_end = 101, 680, 2740, 2909
    c_all, c_aj, c_n, c_av, c_v = [], [], [], [], []

    b_aj_start, b_aj_end = 2, 762
    b_n_start, b_n_end = 977, 3601
    b_av_start, b_av_end = 763, 976
    b_v_start, b_v_end = 3602, 4753
    if limited:
        b_aj_end, b_n_end, b_av_end, b_v_end = 101, 1076, 862, 3701
    b_all, b_aj, b_n, b_av, b_v = [], [], [], [], []

    coca = load_workbook(filename = 'COCA - Noun, Verb, Adj, Adv only - Valid Only.xlsx')
    ws = coca.active

    for i in range (c_aj_start, c_aj_end+1):
        c_aj.append(ws.cell(row=i, column=c_col).value)
    for i in range (c_n_start, c_n_end+1):
        c_n.append(ws.cell(row=i, column=c_col).value)
    for i in range (c_av_start, c_av_end+1):
        c_av.append(ws.cell(row=i, column=c_col).value)
    for i in range (c_v_start, c_v_end+1):
        c_v.append(ws.cell(row=i, column=c_col).value)

    coca = load_workbook(filename = 'COCA - Noun, Verb, Adj, Adv only - Valid Only - Order Sort.xlsx')
    ws = coca.active

    if limited:
        for i in range (2, 502):
            c_all.append(ws.cell(row=i, column=c_col).value)
    else:
        for i in range(2, c_v_end):
            c_all.append(ws.cell(row=i, column=c_col).value)

    bnc = load_workbook(filename = 'BNC - Noun, Verb, Adj, Adv only - Valid Only.xlsx')
    ws = bnc.active

    for i in range (b_aj_start, b_aj_end+1):
        b_aj.append(ws.cell(row=i, column=b_col).value)
    for i in range (b_n_start, b_n_end+1):
        b_n.append(ws.cell(row=i, column=b_col).value)
    for i in range (b_av_start, b_av_end+1):
        b_av.append(ws.cell(row=i, column=b_col).value)
    for i in range (b_v_start, b_v_end+1):
        b_v.append(ws.cell(row=i, column=b_col).value)

    bnc = load_workbook(filename = 'BNC - Noun, Verb, Adj, Adv only - Valid Only - Order Sort.xlsx')
    ws = bnc.active

    if limited:
        for i in range (2, 502):
            b_all.append(ws.cell(row=i, column=b_col).value)
    else:
        for i in range(2, b_v_end):
            b_all.append(ws.cell(row=i, column=b_col).value)

    fig, axes = plt.subplots(nrows=5, ncols=2)
    ax0, ax1, ax2, ax3, ax4, ax5, ax6, ax7, ax8, ax9 = axes.flat

    def t_test(data):
        stat = ttest_1samp(data, 0)[0]
        p = ttest_1samp(data, 0)[1]
        return stat, p

    if limited:
        title_part1 = "Top 500 "
        title_part2 = "Top 100 "
    else:
        title_part1 = "All "
        title_part2 = "All "

    #BNC 5x1
    ax0.hist(b_all, 25, range=(-1,1))
    ax0.axvline(np.mean(b_all), color='r', linestyle='dashed', linewidth=2)
    ax0.axvline(np.median(b_all), color='r', linewidth=2)
    ax0.set_title('BNC (' + title_part1 + 'Words) - T-Stat, P-Value: ' + str(t_test(b_all)))

    ax2.hist(b_aj, 25, range=(-1,1))
    ax2.axvline(np.mean(b_aj), color='r', linestyle='dashed', linewidth=2)
    ax2.axvline(np.median(b_aj), color='r', linewidth=2)
    ax2.set_title('BNC (' + title_part2 + 'Adj) - T-Stat, P-Value: ' + str(t_test(b_aj)))

    ax4.hist(b_n, 25, range=(-1,1))
    ax4.axvline(np.mean(b_n), color='r', linestyle='dashed', linewidth=2)
    ax4.axvline(np.median(b_n), color='r', linewidth=2)
    ax4.set_title('BNC (' + title_part2 + 'Nouns) - T-Stat, P-Value: ' + str(t_test(b_n)))

    ax8.hist(b_av, 25, range=(-1,1))
    ax8.axvline(np.mean(b_av), color='r', linestyle='dashed', linewidth=2)
    ax8.axvline(np.median(b_av), color='r', linewidth=2)
    ax8.set_title('BNC (' + title_part2 + 'Adv) - T-Stat, P-Value: ' + str(t_test(b_av)))

    ax6.hist(b_v, 25, range=(-1,1))
    ax6.axvline(np.mean(b_v), color='r', linestyle='dashed', linewidth=2)
    ax6.axvline(np.median(b_v), color='r', linewidth=2)
    ax6.set_title('BNC (' + title_part2 + 'Verbs) - T-Stat, P-Value: ' + str(t_test(b_v)))

    #COCA 5x1
    ax1.hist(c_all, 25, range=(-1,1))
    ax1.axvline(np.mean(c_all), color='r', linestyle='dashed', linewidth=2)
    ax1.axvline(np.median(c_all), color='r', linewidth=2)
    ax1.set_title('COCA (' + title_part1 + 'Words) - T-Stat, P-Value: ' + str(t_test(c_all)))

    ax3.hist(c_aj, 25, range=(-1,1))
    ax3.axvline(np.mean(c_aj), color='r', linestyle='dashed', linewidth=2)
    ax3.axvline(np.median(c_aj), color='r', linewidth=2)
    ax3.set_title('COCA (' + title_part2 + 'Adj) - T-Stat, P-Value: ' + str(t_test(c_aj)))

    ax5.hist(c_n, 25, range=(-1,1))
    ax5.axvline(np.mean(c_n), color='r', linestyle='dashed', linewidth=2)
    ax5.axvline(np.median(c_n), color='r', linewidth=2)
    ax5.set_title('COCA (' + title_part2 + 'Nouns) - T-Stat, P-Value: ' + str(t_test(c_n)))

    ax9.hist(c_av, 25, range=(-1,1))
    ax9.axvline(np.mean(c_av), color='r', linestyle='dashed', linewidth=2)
    ax9.axvline(np.median(c_av), color='r', linewidth=2)
    ax9.set_title('COCA (' + title_part1 + 'Adv) - T-Stat, P-Value: ' + str(t_test(c_av)))

    ax7.hist(c_v, 25, range=(-1,1))
    ax7.axvline(np.mean(c_v), color='r', linestyle='dashed', linewidth=2)
    ax7.axvline(np.median(c_v), color='r', linewidth=2)
    ax7.set_title('COCA (' + title_part1 + 'Verbs) - T-Stat, P-Value: ' + str(t_test(c_v)))

    fig.subplots_adjust(hspace=.5)
    plt.show()

def write_3_by_2_misc(before, limited, title):

    if before:
        b_col, c_col = 5, 6
    else:
        b_col, c_col = 6, 7

    fig, axes = plt.subplots(nrows=2, ncols=3)
    ax0, ax1, ax2, ax3, ax4, ax5 = axes.flat

    #BNC CONCRETENESS RATINGS
    bnc = load_workbook('BNC - Noun, Verb, Adj, Adv only - Valid Only - With Concrete.xlsx')
    ws = bnc.active
    if limited:
        conc_range = 501
    else:
        conc_range = 4688
    bnc_conc_pearson = [ws.cell(row=i, column=b_col).value for i in range(2, conc_range)]
    bnc_conc = [ws.cell(row=i, column=8).value for i in range(2, conc_range)]

    #BNC FREQUENCY AND EXTERNALITY
    bnc = load_workbook('BNC - Noun, Verb, Adj, Adv only - Valid Only - Order Sort.xlsx')
    ws = bnc.active
    if limited:
        freq_ext_range = 501
    else:
        freq_ext_range = 4754
    bnc_freq = [np.log(int(ws.cell(row=i, column=4).value)) for i in range(2, freq_ext_range)]
    bnc_pearsons = [ws.cell(row=i, column=b_col).value for i in range(2, freq_ext_range)]
    bnc_ext_pearsons = [ws.cell(row=i, column=b_col).value for i in range(2, freq_ext_range) if
                         ws.cell(row=i, column=7).value == 1]
    bnc_int_pearsons = [ws.cell(row=i, column=b_col).value for i in range(2, freq_ext_range) if
                         ws.cell(row=i, column=7).value == 0]

    #COCA CONCRETENESS RATINGS
    coca = load_workbook('COCA - Noun, Verb, Adj, Adv only - Valid Only - With Concrete.xlsx')
    ws = coca.active

    if limited:
        conc_range = 501
    else:
        conc_range = 3689
    coca_conc_pearson = [ws.cell(row=i, column=c_col).value for i in range(2, conc_range)]
    coca_conc = [ws.cell(row=i, column=9).value for i in range(2, conc_range)]

    #COCA FREQUENCY AND EXTERNALITY
    coca = load_workbook('COCA - Noun, Verb, Adj, Adv only - Valid Only - Order Sort.xlsx')
    ws = coca.active
    if limited:
        freq_ext_range = 501
    else:
        freq_ext_range = 3725
    coca_freq = [np.log(ws.cell(row=i, column=4).value) for i in range(2, freq_ext_range)]
    coca_pearsons = [ws.cell(row=i, column=c_col).value for i in range(2, freq_ext_range)]
    coca_ext_pearsons = [ws.cell(row=i, column=c_col).value for i in range(2, freq_ext_range) if
                         ws.cell(row=i, column=8).value == 1]
    coca_int_pearsons = [ws.cell(row=i, column=c_col).value for i in range(2, freq_ext_range) if
                         ws.cell(row=i, column=8).value ==0]

    if limited:
        title_part = "Top 500 Words"
    else:
        title_part = "All Words"

    #Externality Scatterplot

    ext_stat = round(float(ttest_ind(bnc_ext_pearsons, bnc_int_pearsons)[0]),5)
    ext_p = round(float(ttest_ind(bnc_ext_pearsons, bnc_int_pearsons)[1]),5)
    mean_ext, mean_int = round(float(np.mean(bnc_ext_pearsons)), 5), round(float(np.mean(bnc_int_pearsons)), 5)
    ax0.boxplot([bnc_ext_pearsons, bnc_int_pearsons], showmeans=True, labels=["External, mean = " + str(mean_ext), "Non-External, mean = " + str(mean_int)])
    ax0.set_title("BNC (" + title_part + ") - Externality, \n (Stat, 1tP-Value): " + str((ext_stat, ext_p/2)))

    ext_stat = round(float(ttest_ind(coca_ext_pearsons, coca_int_pearsons)[0]),5)
    ext_p = round(float(ttest_ind(coca_ext_pearsons, coca_int_pearsons)[1]),5)
    mean_ext, mean_int = round(np.mean(coca_ext_pearsons), 5), round(np.mean(coca_int_pearsons), 5)
    ax3.boxplot([coca_ext_pearsons, coca_int_pearsons], showmeans=True, labels=["External, mean = " + str(mean_ext), "Non-External, mean = " + str(mean_int)])
    ax3.set_title("COCA (" + title_part + ")  - Externality, \n (Stat, 1tP-Value): " + str((ext_stat, ext_p/2)))

    #Concreteness Scatterplot

    ax1.scatter(bnc_conc, bnc_conc_pearson)
    slope = numpy.polyfit(bnc_conc, bnc_conc_pearson, 1)
    x, y = round(numpy.polyfit(bnc_conc, bnc_conc_pearson, 1)[0], 5), round(numpy.polyfit(bnc_conc, bnc_conc_pearson, 1)[1], 5)
    slope_w = "y = " + str(y) + " + " + str(x) + "x"
    pearson_r, pearson_p = pearsonr(bnc_conc, bnc_conc_pearson)
    pearson_r, pearson_p = round(float(pearson_r), 5), round(float(pearson_p), 5)
    ax1.plot(bnc_conc, numpy.poly1d(slope)(bnc_conc))
    ax1.set_title("BNC (" + title_part + ")  - Concreteness, \n Slope: " + str(slope_w) + "\n (pearsonR, 1tP-Value): " + str((pearson_r, pearson_p)))

    ax4.scatter(coca_conc, coca_conc_pearson)
    slope = numpy.polyfit(coca_conc, coca_conc_pearson, 1)
    x, y = round(numpy.polyfit(coca_conc, coca_conc_pearson, 1)[0], 5), round(numpy.polyfit(coca_conc, coca_conc_pearson, 1)[1], 5)
    slope_w = "y = " + str(y) + " + " + str(x) + "x"
    pearson_r, pearson_p = pearsonr(coca_conc, coca_conc_pearson)
    pearson_r, pearson_p = round(float(pearson_r), 5), round(float(pearson_p), 5)
    ax4.plot(coca_conc, numpy.poly1d(slope)(coca_conc))
    ax4.set_title("COCA (" + title_part + ")  - Concreteness, \n Slope: " + str(slope_w) + "\n (pearsonR, 1tP-Value): " + str((pearson_r, pearson_p)))

    #Log Frequency Scatterplot

    ax2.scatter(bnc_freq, bnc_pearsons)
    slope = numpy.polyfit(bnc_freq, bnc_pearsons, 1)
    x, y = round(numpy.polyfit(bnc_freq, bnc_pearsons, 1)[0], 5), round(numpy.polyfit(bnc_freq, bnc_pearsons, 1)[1], 5)
    slope_w = "y = " + str(y) + " + " + str(x) + "x"
    pearson_r, pearson_p = pearsonr(bnc_freq, bnc_pearsons)
    pearson_r, pearson_p = round(float(pearson_r), 5), round(float(pearson_p), 5)
    ax2.plot(bnc_freq, numpy.poly1d(slope)(bnc_freq))
    ax2.set_title("BNC (" + title_part + ")  - log(Frequency), \n Slope: " + str(slope_w) + "\n (pearsonR, 1tP-Value): " + str((pearson_r, pearson_p)))

    ax5.scatter(coca_freq, coca_pearsons)
    slope = numpy.polyfit(coca_freq, coca_pearsons, 1)
    x, y = round(numpy.polyfit(coca_freq, coca_pearsons, 1)[0], 5), round(numpy.polyfit(coca_freq, coca_pearsons, 1)[1], 5)
    slope_w = "y = " + str(y) + " + " + str(x) + "x"
    pearson_r, pearson_p = pearsonr(coca_freq, coca_pearsons)
    pearson_r, pearson_p = round(float(pearson_r), 5), round(float(pearson_p), 5)
    ax5.plot(coca_freq, numpy.poly1d(slope)(coca_freq))
    ax5.set_title("COCA (" + title_part + ")  - log(Frequency), \n Slope: " + str(slope_w) + "\n (pearsonR, 1tP-Value): " + str((pearson_r, pearson_p)))

    fig.subplots_adjust(hspace=.5)
    plt.show()

def write_scatterplot(x, y, axis, title):
    axis.scatter(x, y)
    slope = np.polyfit(x, y, 1)
    lx, ly = round(np.polyfit(x, y, 1)[0], 5) , round(np.polyfit(x, y, 1)[1], 5)
    slope_str = "y = " + str(ly) + " + " + str(lx) + "x"
    pearson_r, pearson_p = pearsonr(x, y)
    pearson_r, pearson_p = round(float(pearson_r), 5), round(float(pearson_p), 5)
    pearson_str = "Pearson(r, p) : (" + str(pearson_r) + ", " + str(pearson_p) + ")"
    axis.plot(x, numpy.poly1d(slope)(x))
    axis.set_title(title + "\n" + slope_str + "\n" + pearson_str)

