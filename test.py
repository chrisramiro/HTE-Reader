from word_data_structures import *
from scipy.sparse import csr_matrix
from scipy.sparse.csgraph import minimum_spanning_tree
from excel import write_sense_matrix
from word import *
from word_data_structures import *
from openpyxl import Workbook, load_workbook
import os
import bs4
import requests

corpus = load_workbook('BNC - Noun, Verb, Adj, Adv only - Valid Only - Order Sort.xlsx')
ws = corpus.active
words = []

for i in range(2950, 4754):
    print("I'm crying: " + str(i) + " " + ws.cell(row=i, column=2).value, flush=True)
    word = Word(ws.cell(row=i, column=2).value, "all")
    if word.viable:
        ws.cell(row=i, column=10).value = word.pearson_avg_after
        ws.cell(row=i, column=11).value = word.avg_externality
        ws.cell(row=i, column=12).value = word.num_nb_senses
    else:
        ws.cell(row=i, column=10).value = "Not Viable"
    if i % 50 == 0:
        corpus.save('BNC - Noun, Verb, Adj, Adv only - Valid Only - Order Sort.xlsx')


corpus.save('BNC - Noun, Verb, Adj, Adv only - Valid Only - Order Sort.xlsx')

"""
wordy = Word("leg", "all")

print(create_mst(wordy))

print("REAL:      " + str(create_real(wordy)))

print("NULL:      " + str(create_null(wordy)))

print("PROTOTYPE: " + str(create_prototype(wordy)))

print("EXEMPLAR:  " + str(create_exemplar(wordy)))

print("SIMPLE:    " + str(create_simple_chain(wordy)))


 YI-WEI'S WORD LIST PULLER

bnc = load_workbook('yiwei_wordlist.xlsx')
ws = bnc.active
words = []

for i in range(1,185):
    words.append(ws.cell(row=i, column=1).value)

workbook = Workbook()
ws1 = workbook.active

worksheets = {}

for i in range(len(words)):
    worksheets["ws" + str(i)] = workbook.create_sheet(title=words[i])

for j in range(len(words)):
    print("word" + str(j), flush=True)
    word = Word(words[j], "all")
    ws1 = worksheets["ws" + str(j)]
    ws1['A1'], ws1['B1'], ws1['C1'], ws1['D1'], ws1[
        'E1'] = "Starting Date", "Word Form", "Identifiers", "Categories", "PoS"
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 110
    ws1.column_dimensions['D'].width = 35
    for num in range(len(word.senses)):
        ws1.cell(row=num + 2, column=1).value = word.senses[num].date
        ws1.cell(row=num + 2, column=2).value = word.senses[num].word_form
        ws1.cell(row=num + 2, column=3).value = word.senses[num].identifiers
        ws1.cell(row=num + 2, column=4).value = word.senses[num].categories
        ws1.cell(row=num + 2, column=5).value = word.senses[num].PoS

save_loc = os.getcwd() + "\\Sense Lists\\" + "[""] " + word.label + ".xlsx"
workbook.save(save_loc)

"""