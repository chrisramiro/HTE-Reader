from source import *
from openpyxl import Workbook

input_word, word_list = word_list_prompt()

wb = Workbook()
dest_filename = "multi_compared_" + input_word + ".xlsx"
ws1 = wb.active
length = len(word_list)
abc = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

for num in range(2, length + 2):
	row1, col1 = ws1.cell(row = 1, column = num), ws1.cell(row = num, column = 1)
	row1.value = ''.join(str(word_list[num - 2]))
	col1.value = ''.join(str(word_list[num - 2]))

categories = []
for word in word_list:
	categories.append(categorizer(word[-1]))

for row in range(2, length +2):
	for col in range(2, length + 2):
		cur_cell = ws1.cell(row = row, column = col)
		#when a sense is matched with itself, its difference value is maximal
		if row == col:
			cur_cell.value = "max"
		else:
			similarity_score, index = 0, 0
			list_range = min(len(categories[row -2]), len(categories[col - 2]))
			for r in range(0, list_range):
				if categories[row - 2][r] == categories[col - 2][r]:
					similarity_score += 1
				else:
					break
			cur_cell.value = similarity_score

wb.save(filename = dest_filename)
