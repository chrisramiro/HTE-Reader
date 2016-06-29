from source import *
from openpyxl import Workbook
import requests
from bs4 import *
import sys

print('Enter a string of all the words that fit into a certain part-of-speech category: ')
rqst_words = input('(For example: "bear dog cat porpoise anteater emu pangolin") ')
print('Enter your desired part of speech ("aj", "av", "cj", "in", "n", "p", "ph", "v")')
input_PoS = input('Replying with "all" anything else will return all: ')
file_name = input('Please name this file (sans file extension): ')
dest_filename = file_name + ".xlsx"

word_list = rqst_words.split()
worksheets = {}
wb = Workbook()
p_rows = []

for i in range(len(word_list)):
    worksheets["ws" + str(i)] = wb.create_sheet(title = word_list[i])

pearson_ws = wb.create_sheet(title = "Pearson")
for j in range(len(word_list)): #aka for j in len(worksheets)
    defunct_word = False
    sense_list = make_word_list(word_list[j], input_PoS)
    current_ws = worksheets["ws" + str(j)]

    index = 0
    while True:
        try:
            if sense_list[index][0] != sense_list[index + 1][0]:
                break
        except:
            defunct_word = True #implies that there are no words to compare against the base senses
            current_ws['A350'] = "defunct word"
            break
        index += 1

    pearson_row = 2 * (index + 3)
    p_rows.append(pearson_row)

    if defunct_word:
        continue

    x_vals, y_vals = make_averaged(word_list[j], sense_list, index, current_ws)

    current_ws.cell(row = pearson_row, column = 1 ).value = "PEARSON(y,x)"
    current_ws.cell(row = pearson_row - 1, column = 1).value = "RANGE(x,y)"
    current_ws.cell(row = pearson_row - 1, column = 2).value = x_vals
    current_ws.cell(row = pearson_row - 1, column = 3).value = y_vals
    current_ws.cell(row = pearson_row, column = 2 ).value = "=PEARSON(" + word_list[j] + "!" + y_vals + "," + word_list[j] + "!" + x_vals + ")"

pearson_ws['A1'] = "Words"
pearson_ws['B1'] = "Correlations"
pearson_ws['C1'] = "Bins"
bins = [-.8, -.6, -.4, -.2, 0, .2, .4, .6, .8, 1]
for k in range(len(word_list)):
    current_ws = worksheets["ws" + str(k)]
    if current_ws['A350'] == "defunct word":
        continue
    xp = current_ws.cell(row = p_rows[k] - 1, column = 2).value
    yp = current_ws.cell(row = p_rows[k] - 1, column = 3).value
    pearson_ws['A' + str(k + 2)] = word_list[k]
    pearson_ws['B' + str(k + 2)] = "=PEARSON(" + word_list[k] + "!" + str(yp) + "," + word_list[k] + "!" + str(xp) + ")"

for l in range(len(bins)):
    pearson_ws['C' + str(l + 2)] = bins[l]

wb.save(filename = dest_filename)
